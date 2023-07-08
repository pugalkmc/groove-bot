import os

import openpyxl
import requests
from binance.client import Client
from telegram.ext import CommandHandler, MessageHandler, filters, Application
from edit_task import *
from settings_con import *
from task_create import *


def get_ip_address():
    url = "https://checkip.amazonaws.com/"
    payload = {}
    headers = {}
    response = requests.request("GET", url, headers=headers, data=payload)
    return response.text


client = Client(api_key, api_secret)


async def binance_pay(update, context, amount):
    chat_id = update.message.chat_id
    response = client.withdraw(
        coin='TUSD',
        address='0x6B518FD3dF2d4B97BAfE551460dee5A7852df0d6',
        amount=amount,
        network='BSC')
    await bot.send_message(chat_id=chat_id, text=response)


async def start(update, context):
    message = update.message
    chat_id = message.chat_id
    username = message.from_user.username
    if message.chat.type in ['group', 'supergroup']:
        return
    current_time = time_fun.now().strftime("%d-%m-%Y")
    find_people = db.reference(f'peoples/{chat_id}').get() or {}
    if username is None:
        await bot.send_message(chat_id=chat_id, text=f"Hello @{username}\n"
                                                     f"Please set your telegram username in settings!\n"
                                                     f"Check out this document for guidance: Not set")
    elif find_people is None or len(find_people) == 0:
        db.reference(f'peoples/{chat_id}').set({
            'username': username,
            'chat_id': chat_id,
            'first_started': current_time
        })
        await bot.send_message(chat_id=chat_id, text=f"Hello! welcome @{username}")
    else:
        await bot.send_message(chat_id=chat_id, text="Hi! welcome back")
    await menu_button(update, context)


async def cancel(update, context):
    message = update.message
    chat_id = message.chat_id
    if message.chat.type == 'group' or message.chat.type == 'supergroup':
        return
    reply_keyboard = [["settings", "My task"]]
    await bot.send_message(chat_id=chat_id, text="Use menu buttons for quick access",
                           reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True,
                                                            one_time_keyboard=True),
                           reply_to_message_id=update.message.message_id)


async def private_message_handler(update, context):
    message = update.message
    chat_id = message.chat_id
    text = message.text.lower()
    username = update.message.chat.username.lower()
    if username is None:
        await bot.send_message(chat_id=chat_id, text=f"Hello @{username}\n"
                                                     f"Please set your telegram username in settings!\n"
                                                     f"Check out this document for guidance: Not set")
        return
    if text == "cancel":
        await cancel(update, context)
    if "check " in text:
        check = text.split(" ")[1]
        chat = bot.get_chat(check)
        print(chat)
        for i in chat:
            print(i)
        chat_id = chat.id
        print(chat_id)
    if "ip address" == text:
        ip_address = get_ip_address()
        await bot.send_message(chat_id=chat_id, text=ip_address)
    elif "send otp" == text:
        otp = await otp_sender(update, context)
        await bot.send_message(chat_id=chat_id, text=otp)
    elif "pay " in text:
        details_pay = text.split(" ")
        await functions.binance_pay(details_pay[0], details_pay[1], details_pay[3], chat_id)
    elif "pay now " in text:
        text = text.replace("pay now ", "")
        if float(text) < 50:
            await binance_pay(update, context, float(text))
        else:
            await bot.send_message(chat_id=chat_id, text="Suspicious pay")

    elif "my task" == text:
        user_task_list = db.reference(f"tasks").get() or {}
        text = ""
        for i in user_task_list:
            each_task = db.reference(f"tasks/{i}").get() or {}
            if username in each_task["workers"]:
                text += f"Title : {each_task['title']}\nCommand: <code>sheet {each_task['task_id']}</code>\n\n"
        if len(text) <= 1:
            await bot.send_message(chat_id=chat_id, text="You are not assigned for any task right now!")
        else:
            await bot.send_message(chat_id=chat_id, text=f"Task List:\n\n"
                                                         f"{text}"
                                                         f"Just click to copy the command and send it here",
                                   parse_mode="html")
    elif "sheet " in text:
        task_id = text.split(" ")
        if not task_id[1].isnumeric():
            await bot.send_message(chat_id=chat_id, text=f"task id must be number\n")
        elif len(task_id) == 2:
            await sheet_file.spreadsheet(chat_id, task_id[1])
        else:
            await sheet_file.spreadsheet(chat_id, task_id[1], task_id[2])
    elif "payment data" == text:
        get = db.reference(f"peoples").get() or {}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws['A1'] = 'Username'
        ws['B1'] = 'Binance'
        ws['C1'] = 'UPI'
        row = 2
        for data in get.values():
            if 'username' not in data:
                continue
            username = data['username']
            ws.cell(row=row, column=1).value = username
            ws.cell(row=row, column=2).value = data['binance'] if 'binance' in data else "Not set"
            ws.cell(row=row, column=3).value = data['upi'] if 'upi' in data else "Not set"
            row += 1
        wb.save(f"peoples_data.xlsx")
        await bot.send_document(chat_id=chat_id, document=open(f"peoples_data.xlsx", "rb"))
        wb.close()
        os.remove("peoples_data.xlsx")

    elif "pause " in text:
        task_id = text.replace("pause ", "")
        await task_status_switch(chat_id, task_id, option="pause")

    elif "active " in text:
        task_id = text.replace("active ", "")
        await task_status_switch(chat_id, task_id, option="active")


async def group_message_handler(update, context):
    message = update.message
    text = message.text
    username = message.from_user.username
    message_id = message.message_id
    group_id = message.chat.id

    collection_name = time_fun.now().strftime("%d-%m-%Y")
    message_date_ist = time_fun.now().strftime("%H:%M:%S")
    task = db.reference(f'tasks/{group_id}').get() or {}
    if len(task) <= 0 or task['status'] == 'paused':
        return
    inserting = "link" if task['task_type'] == "twitter" else "text"
    if task['task_type'] == 'twitter':
        if 'twitter.com' not in text or len(text) < 15:
            return
        # else:
        #     if "?" in text:
        #         text = text.split("?")[0]
        #         print(text)
        #     messages = db.reference(f"tasks/{task['group_id']}/collection/{collection_name}").get() or {}
        # link_exists = any(message.get('link') == text for message in messages.values())
        # if link_exists:
        #     await bot.send_message(chat_id=chat_id, text="Link already submitted")
        #     return
    if task['task_type'] == 'telegram':
        if not await verify_membership(update, context):
            return

    db.reference(f"tasks/{task['group_id']}/collection/{collection_name}/{message_id}").set({
        'username': username,
        inserting: text,
        'message_id': message_id,
        'chat_id': message.from_user.id,
        'time': message_date_ist
    })


async def task_status_switch(chat_id, task_id, option):
    if task_id.isnumeric():
        get = db.reference(f'task_ids/{task_id}').get() or {}
        if get is None:
            await bot.send_message(chat_id=chat_id, text="Invalid task ID")
            return
        db.reference(f"tasks/{get['group_id']}").update({
            'status': option
        })
        await bot.send_message(chat_id=chat_id, text=f"{get['group_id']} : Task paused!")
        await bot.send_message(chat_id=get['group_id'], text="Task paused")
    else:
        await bot.send_message(chat_id=chat_id, text="Task id must be a number")


async def verify_membership(update, context):
    user_id = update.message.from_user.id
    try:
        member = await bot.get_chat_member(-1001640271166, user_id)
        if member.status in ['member', 'creator', 'administrator']:
            return True
        else:
            return False
    except Exception as e:
        return False


def main():
    dp = Application.builder().token(BOT_TOKEN).build()
    create = ConversationHandler(
        entry_points=[CommandHandler('create_task', create_task)],
        states={
            TITLE: [MessageHandler(filters.TEXT, title)],
            TASK_TYPE: [MessageHandler(filters.TEXT, task_type)],
            CHAT_ID: [MessageHandler(filters.TEXT, chat_id)],
            LIMIT: [MessageHandler(filters.TEXT, limit)],
            MEMBERS_LIST: [MessageHandler(filters.TEXT, members_list)],
            CONFIRM: [MessageHandler(filters.TEXT, confirm)]
        }, fallbacks=[]
    )

    twitter_settings = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^twitter$'), twitter_ids)],
        states={
            TWITTER_UPDATE: [MessageHandler(filters.TEXT, twitter_update)],
            TWITTER_UPDATE_LIST: [MessageHandler(filters.TEXT, twitter_update_list)],
            TWITTER_UPDATE_CONFIRM: [MessageHandler(filters.TEXT, twitter_update_confirm)],
        }, fallbacks=[]
    )

    binance = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^binance$'), binance_start)],
        states={
            BINANCE_OPTIONS: [MessageHandler(filters.TEXT, binance_option)],
            SET_BINANCE: [MessageHandler(filters.TEXT, set_binance)],
        }, fallbacks=[]
    )

    discord = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^discord$'), discord_start)],
        states={
            DISCORD_OPTIONS: [MessageHandler(filters.TEXT, discord_option)],
            SET_DISCORD: [MessageHandler(filters.TEXT, set_discord)],
        }, fallbacks=[]
    )

    upi = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^UPI ID$'), upi_start)],
        states={
            UPI_OPTIONS: [MessageHandler(filters.TEXT, upi_option)],
            SET_UPI: [MessageHandler(filters.TEXT, set_upi)],
        }, fallbacks=[]
    )

    edit_task_con = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^Admin Mode$'), task_id)],
        states={
            ADMIN_MODE: [MessageHandler(filters.TEXT, admin_mode)],
            MEMBER_START: [MessageHandler(filters.TEXT, member_start)],
            SHEET_OPTION: [MessageHandler(filters.TEXT, sheet_option)],
            DATE_RANGE: [MessageHandler(filters.TEXT, date_range)],
            PROCEED_PAYMENT: [MessageHandler(filters.TEXT, proceed_payment)],
            PAYMENT_FOR_TASK: [MessageHandler(filters.TEXT, payment_for_task)],
            CHECK_PASSWORD: [MessageHandler(filters.TEXT, check_password)],
        }, fallbacks=[]
    )
    dp.add_handler(edit_task_con)
    dp.add_handler(upi)
    dp.add_handler(create)
    dp.add_handler(binance)
    dp.add_handler(discord)
    dp.add_handler(twitter_settings)
    dp.add_handler(CommandHandler("cancel", cancel))
    dp.add_handler(MessageHandler(filters.Regex('^settings$'), settings))
    dp.add_handler(CommandHandler("start", cancel))
    dp.add_handler(MessageHandler(filters.ChatType.PRIVATE, private_message_handler))
    dp.add_handler(MessageHandler(filters.ChatType.GROUPS, group_message_handler))
    dp.run_polling()


main()
