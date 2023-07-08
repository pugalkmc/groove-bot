import os

import openpyxl
from firebase_admin import db
from telegram.ext import ConversationHandler

from functions import *


async def spreadsheet(chat_id, task_id=None, date=None):
    if date is None:
        collection_name = datetime.datetime.now().strftime("%d-%m-%Y")
    else:
        collection_name = date
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write the headers
    datas = db.reference(f"task_ids/{task_id}").get() or {}
    if datas is None:
        await bot.send_message(chat_id=chat_id, text="Task id not valid!")
        return 0
    group_id = datas['group_id']
    task_type = datas['task_type']
    task_details = db.reference(f"tasks/{group_id}/collection/{collection_name}").get() or {}
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18

    ws['F1'] = 'Username'
    ws['G1'] = 'Count'
    flex_title = "Message text" if task_type == "telegram" else "tweet link"

    ws['A1'] = 'Username'
    ws['B1'] = flex_title
    ws['C1'] = 'IST Time'

    get = "text" if task_type == "telegram" else "link"
    username_counts = {}
    if task_details is not None and isinstance(task_details, dict):
        row = 2
        for task_id, task_info in task_details.items():
            if task_info is not None and isinstance(task_info, dict):
                username = task_info.get('username')
                text = task_info.get(get)
                time = task_info.get('time')
                if username is not None:
                    ws.cell(row=row, column=1).value = username
                if text is not None:
                    ws.cell(row=row, column=2).value = text
                if time is not None:
                    ws.cell(row=row, column=3).value = time

                if username in username_counts:
                    username_counts[username]['count'] += 1
                else:
                    username_counts[username] = {'count': 1}
                row += 1
    row = 1  # Assuming the initial row is 1
    for username, data in username_counts.items():
        count = data['count']
        ws.cell(row=row, column=6).value = username
        ws.cell(row=row, column=7).value = count
        row += 1

    wb.save(f"{collection_name}.xlsx")
    await bot.send_document(chat_id=chat_id, document=open(f"{collection_name}.xlsx", "rb"))
    wb.close()
    os.remove(f"{collection_name}.xlsx")
